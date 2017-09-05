# _*_ coding:utf-8 _*_
# 用python读取excel文件

from openpyxl import load_workbook

# 打开一个workbook
wb = load_workbook(filename="/Users/huixing/desktop/jiance_list.xlsx")

# 获取当前活跃的worksheet,默认就是第一个worksheet
# ws = wb.active

# 当然也可以使用下面的方法

# 获取所有表格(worksheet)的名字
sheets = wb.get_sheet_names()
# 第一个表格的名称
sheet_first = sheets[0]
# 获取特定的worksheet
ws = wb.get_sheet_by_name(sheet_first)

# 获取表格所有行和列，两者都是可迭代的
rows = ws.rows
columns = ws.columns

# 迭代所有的行
for row in rows:
    line = [col.value for col in row]
    print(line[0],line[1]+line[2]+"损伤",line[3])

# for column in columns:
#     line2 = [col.value for col in column]
#     print(line2)

# 通过坐标读取值
# print(ws.cell('B5').value)  # A表示列,1表示行
# print(ws.cell(row=70, column=2).value)